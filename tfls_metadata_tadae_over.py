# -*- coding: utf-8 -*-
"""
生成不良事件汇总表 metadata t14_3_1-1_1.xlsx。

从 edcdef_code.sas7bdat 解析 AE 码表（AESEV/AETOXGR、AEACN）及 EDC 变量存在性（AEDIS、AESI），
输入：PATH_T14_3_1_1_1（输出 xlsx）、PATH_EDCDEF_CODE（edcdef_code.sas7bdat 或 .xlsx）。
依赖：pandas, openpyxl, pyreadstat（读 sas7bdat）。可直接运行或调用 run_t14_3_1_1_1_init(path_t14, path_edc_code)。
"""
import logging
import os
import shutil
import sys
from datetime import datetime

logger = logging.getLogger(__name__)

# ---------- 输入宏参数 ----------
PATH_T14_3_1_1_1 = ""   # t14_3_1-1_1.xlsx 输出路径
PATH_EDCDEF_CODE = ""   # edcdef_code.sas7bdat 


def _setup_log_file():
    """将本模块日志同时输出到本地文件（logs/tfls_metadata_tadae_over.log），仅添加一次。"""
    if any(isinstance(h, logging.FileHandler) for h in logger.handlers):
        return
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
    try:
        os.makedirs(log_dir, exist_ok=True)
        log_path = os.path.join(log_dir, "tfls_metadata_tadae_over.log")
        fh = logging.FileHandler(log_path, mode="a", encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
        logger.addHandler(fh)
        logger.setLevel(logging.DEBUG)
    except Exception as e:
        logger.debug("无法创建日志文件 %s: %s", log_dir, e)


_setup_log_file()


_EDC_CODE_GRP_AE = "AE"
_CODE_NAMES_AESEV = ("AESEV", "AETOXGR")   # AESEV 的值：来自 CODE_NAME IN ('AESEV','AETOXGR')
_AEACN_EXCLUDE_RAW = ("NCHANGE", "NA")
_PLACEHOLDER_AESEV = "[AESEV]"
_PLACEHOLDER_AEACN = "[AEACN]"

# 表头与 t14_3_1-1_1-final.xlsx 一致，程序硬写
_T14_3_1_1_1_HEADER = ("ROW", "TEXT", "MASK", "LINE_BREAK", "INDENT", "FILTER")

# 第 2、4 行 FILTER 模板（占位符 [AESEV] 替换为实际 CODE_LABEL，SAS 中需引号故替换为 'label'）
_ROW2_FILTER_TEMPLATE = 'trtemfl="Y" and aesev = [AESEV]'
_ROW4_FILTER_TEMPLATE = 'traefl ="Y" and aesev = [AESEV]'
# 第 5、6 行 TEXT/FILTER 模板（占位符 [AEACN] 替换为实际 CODE_LABEL）
_ROW5_TEXT_TEMPLATE = "导致[AEACN]的TEAE"
_ROW5_FILTER_TEMPLATE = 'trtemfl="Y" and aeacn = [AEACN]'
_ROW6_TEXT_TEMPLATE = "导致[AEACN]的TRAE"
_ROW6_FILTER_TEMPLATE = 'traefl ="Y" and aeacn = [AEACN]'

# 除第 2、4、5、6 行外，其他行直接赋值（ROW 列由程序按顺序生成）。INDENT 为整型，无缩进为 None。
_FIXED_ROWS = (
    {"TEXT": "治疗期间出现的不良事件 (TEAE)", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'trtemfl="Y"'},
    {"TEXT": "治疗相关不良事件 (TRAE)", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'traefl ="Y"'},
    {"TEXT": "导致退出研究的TEAE", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'trtemfl="Y" and prxmatch("/^(是|y|yes)\\s*$/i", aedis)'},
    {"TEXT": "导致退出研究的TRAE", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'traefl ="Y" and prxmatch("/^(是|y|yes)\\s*$/i", aedis)'},
    {"TEXT": "治疗期间出现的严重不良事件(SAE)", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'trtemfl="Y" and prxmatch("/^(是|y|yes)\\s*$/i", aeser)'},
    {"TEXT": "治疗相关 SAE", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'traefl ="Y" and prxmatch("/^(是|y|yes)\\s*$/i", aeser)'},
    {"TEXT": "导致死亡的TEAE", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'trtemfl="Y" and (prxmatch("/^(死亡|death)\\s*$/i", aeout) or prxmatch("/^(是|y|yes)\\s*$/i", aesdth))'},
    {"TEXT": "导致死亡的TRAE", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'traefl ="Y" and (prxmatch("/^(死亡|death)\\s*$/i", aeout) or prxmatch("/^(是|y|yes)\\s*$/i", aesdth))'},
    {"TEXT": "治疗期间出现的特别关注的不良事件(AESI)", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'trtemfl="Y" and prxmatch("/^(是|y|yes)\\s*$/i", aesi)'},
    {"TEXT": "治疗相关AESI", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": 'traefl ="Y" and prxmatch("/^(是|y|yes)\\s*$/i", aesi)'},
)
# _FIXED_ROWS 对应顺序：行1(索引0)、行3(索引1)、行7(2)、行8(3)、行9(4)、行10(5)、行11(6)、行12(7)、行13(8)、行14(9)


def _find_excel_column(df, candidates):
    """在 DataFrame 列名中查找匹配项（忽略大小写、首尾空格）。返回列名或 None。"""
    cols = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = cand.strip().lower()
        for col_key, col_orig in cols.items():
            if col_key == k:
                return col_orig
        for col_key, col_orig in cols.items():
            if k in col_key and len(col_key) >= len(k):
                return col_orig
            if col_key in k and len(col_key) >= len(k):
                return col_orig
    return None


def read_edcdef_code_ae(edc_path):
    """
    读取 edcdef_code（.sas7bdat ），仅保留 CODE_GRP='AE'，解析：
    - aesev_values: CODE_NAME IN ('AESEV','AETOXGR') 的 CODE_LABEL 列表，按 CODE_ORDER 排序
    - aeacn_values: CODE_NAME='AEACN' 且 CODE_RAW_VALUE not in ('NCHANGE','NA') 的 CODE_LABEL 列表，按 CODE_ORDER 排序
    - aedis_exist: 'Y' 若存在 EDC_VARIABLE='AEDIS'，否则 'N'
    - aesi_exist: 'Y' 若存在 EDC_VARIABLE='AESI'，否则 'N'
    返回 dict 含键 aesev_values, aeacn_values, aedis_exist, aesi_exist。
    """
    try:
        import pandas as pd
    except ImportError:
        raise RuntimeError("请先安装 pandas：pip install pandas")

    if not edc_path or not os.path.isfile(edc_path):
        return {"aesev_values": [], "aeacn_values": [], "aedis_exist": "N", "aesi_exist": "N"}

    ext = os.path.splitext(edc_path)[1].lower()
    if ext == ".sas7bdat":
        try:
            import pyreadstat
            df, _ = pyreadstat.read_sas7bdat(edc_path)
        except ImportError:
            raise RuntimeError("读取 SAS 数据集需要 pyreadstat：pip install pyreadstat")
    elif ext in (".xlsx", ".xls"):
        df = pd.read_excel(edc_path, header=0)
    else:
        return {"aesev_values": [], "aeacn_values": [], "aedis_exist": "N", "aesi_exist": "N"}

    if df is None or df.empty:
        return {"aesev_values": [], "aeacn_values": [], "aedis_exist": "N", "aesi_exist": "N"}

    col_grp = _find_excel_column(df, ("CODE_GRP", "Code_Grp", "code_grp"))
    col_name = _find_excel_column(df, ("CODE_NAME", "Code_Name", "code_name"))
    col_order = _find_excel_column(df, ("CODE_ORDER", "Code_Order", "code_order", "CODE_ORDER_R", "Code_Order_R"))
    col_label = _find_excel_column(df, ("CODE_LABEL", "Code_Label", "code_label"))
    col_raw = _find_excel_column(df, ("CODE_RAW_VALUE", "Code_Raw_Value", "code_raw_value"))
    col_edc_var = _find_excel_column(df, ("EDC_VARIABLE", "Edc_Variable", "edc_variable"))

    if col_grp is None or col_label is None or col_name is None:
        return {"aesev_values": [], "aeacn_values": [], "aedis_exist": "N", "aesi_exist": "N"}

    grp_col = df[col_grp].astype(str).str.strip().str.upper()
    ae_mask = grp_col == _EDC_CODE_GRP_AE
    ae_df = df.loc[ae_mask]
    if ae_df.empty:
        return {"aesev_values": [], "aeacn_values": [], "aedis_exist": "N", "aesi_exist": "N"}

    # AESEV 的值：CODE_NAME IN ('AESEV','AETOXGR')，按 CODE_ORDER 排序
    aesev_tuples = []
    for _, row in ae_df.iterrows():
        name = str(row.get(col_name, "") or "").strip().upper()
        if name not in (c.upper() for c in _CODE_NAMES_AESEV):
            continue
        label = str(row.get(col_label, "") or "").strip()
        if not label:
            continue
        order_val = row.get(col_order) if col_order else 0
        try:
            order_val = float(order_val) if order_val is not None and str(order_val).strip() else 0
        except (ValueError, TypeError):
            order_val = 0
        aesev_tuples.append((order_val, label))
    aesev_tuples.sort(key=lambda x: x[0])
    aesev_values = [lb for _, lb in aesev_tuples]

    # AEACN：CODE_NAME='AEACN' 且 CODE_RAW_VALUE not in ('NCHANGE','NA')，按 CODE_ORDER 排序
    aeacn_tuples = []
    for _, row in ae_df.iterrows():
        name = str(row.get(col_name, "") or "").strip().upper()
        if name != "AEACN":
            continue
        raw_val = str(row.get(col_raw, "") or "").strip().upper() if col_raw else ""
        if raw_val in (x.upper() for x in _AEACN_EXCLUDE_RAW):
            continue
        label = str(row.get(col_label, "") or "").strip()
        if not label:
            continue
        order_val = row.get(col_order) if col_order else 0
        try:
            order_val = float(order_val) if order_val is not None and str(order_val).strip() else 0
        except (ValueError, TypeError):
            order_val = 0
        aeacn_tuples.append((order_val, label))
    aeacn_tuples.sort(key=lambda x: x[0])
    aeacn_values = [lb for _, lb in aeacn_tuples]

    # AEDIS_EXIST / AESI_EXIST
    aedis_exist = "N"
    aesi_exist = "N"
    if col_edc_var and col_edc_var in ae_df.columns:
        for _, row in ae_df.iterrows():
            edc_val = str(row.get(col_edc_var, "") or "").strip().upper()
            if edc_val == "AEDIS":
                aedis_exist = "Y"
            if edc_val == "AESI":
                aesi_exist = "Y"

    return {
        "aesev_values": aesev_values,
        "aeacn_values": aeacn_values,
        "aedis_exist": aedis_exist,
        "aesi_exist": aesi_exist,
    }


def _quoted_sas_value(label):
    """SAS 字符串值：单引号包裹，内部单引号双写。"""
    s = (label or "").replace("'", "''")
    return "'" + s + "'"


def build_t14_3_1_1_1_rows(ae_data):
    """
    按 t14_3_1-1_1-final.xlsx 逻辑构建行列表。
    - 第 1、3、7～14 行：固定赋值。
    - 第 2、4 行：按 AESEV 的值（CODE_NAME IN ('AESEV','AETOXGR')）扩展，TEXT/FILTER 中 [AESEV] 替换为 CODE_LABEL。
    - 第 5、6 行：按 AEACN 的值扩展，TEXT/FILTER 中 [AEACN] 替换为 CODE_LABEL。
    返回 list of dict，每 dict 含 ROW, TEXT, MASK, LINE_BREAK, INDENT, FILTER。
    """
    rows = []
    defaults = {"ROW": "", "TEXT": "", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": ""}
    aesev_values = ae_data.get("aesev_values", [])
    aeacn_values = ae_data.get("aeacn_values", [])

    row_num = 1

    # 第 1 行（固定）
    rows.append({**defaults, "ROW": row_num, **_FIXED_ROWS[0]})
    row_num += 1

    # 第 2 行：按 AESEV 扩展，TEXT 和 FILTER 中 [AESEV] 替换为 CODE_LABEL
    for label in aesev_values:
        q = _quoted_sas_value(label)
        rows.append({
            **defaults,
            "ROW": row_num,
            "TEXT": label,
            "MASK": "",
            "LINE_BREAK": "",
            "INDENT": 1,
            "FILTER": _ROW2_FILTER_TEMPLATE.replace(_PLACEHOLDER_AESEV, q),
        })
        row_num += 1

    # 第 3 行（固定）
    rows.append({**defaults, "ROW": row_num, **_FIXED_ROWS[1]})
    row_num += 1

    # 第 4 行：按 AESEV 扩展
    for label in aesev_values:
        q = _quoted_sas_value(label)
        rows.append({
            **defaults,
            "ROW": row_num,
            "TEXT": label,
            "MASK": "",
            "LINE_BREAK": "",
            "INDENT": 1,
            "FILTER": _ROW4_FILTER_TEMPLATE.replace(_PLACEHOLDER_AESEV, q),
        })
        row_num += 1

    # 第 5、6 行：按 AEACN 扩展，同一 AEACN 先输出「导致[AEACN]的TEAE」再输出「导致[AEACN]的TRAE」
    for label in aeacn_values:
        q = _quoted_sas_value(label)
        lbl = label or ""
        rows.append({
            **defaults,
            "ROW": row_num,
            "TEXT": _ROW5_TEXT_TEMPLATE.replace(_PLACEHOLDER_AEACN, lbl),
            "MASK": "",
            "LINE_BREAK": "",
            "INDENT": None,
            "FILTER": _ROW5_FILTER_TEMPLATE.replace(_PLACEHOLDER_AEACN, q),
        })
        row_num += 1
        rows.append({
            **defaults,
            "ROW": row_num,
            "TEXT": _ROW6_TEXT_TEMPLATE.replace(_PLACEHOLDER_AEACN, lbl),
            "MASK": "",
            "LINE_BREAK": "",
            "INDENT": None,
            "FILTER": _ROW6_FILTER_TEMPLATE.replace(_PLACEHOLDER_AEACN, q),
        })
        row_num += 1

    # 第 7～14 行（固定）
    for fixed in _FIXED_ROWS[2:]:
        rows.append({**defaults, "ROW": row_num, **fixed})
        row_num += 1

    return rows


def _cell_display_width(val):
    """估算单元格显示宽度（中文约 2 单位，英文/数字 1 单位）。"""
    if val is None:
        return 0
    s = str(val)
    return sum(2 if "\u4e00" <= c <= "\u9fff" else 1 for c in s)


def write_t14_3_1_1_1_xlsx(xlsx_path, rows):
    """将 t14_3_1-1_1 行写入 Excel，表头：ROW, TEXT, MASK, LINE_BREAK, INDENT, FILTER。INDENT 列为整数型，列宽按内容调整。"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "不良事件"
    ws.append(list(_T14_3_1_1_1_HEADER))
    defaults = {"ROW": "", "TEXT": "", "MASK": "", "LINE_BREAK": "", "INDENT": None, "FILTER": ""}
    for r in rows:
        row = [r.get(k, defaults.get(k, "")) for k in _T14_3_1_1_1_HEADER]
        ws.append(row)
    # 按内容调整列宽
    ncols = len(_T14_3_1_1_1_HEADER)
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        max_w = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            w = _cell_display_width(cell.value)
            if w > max_w:
                max_w = w
        # 加少量余量，并限制最大宽度
        width = min(max_w + 1, 80)
        ws.column_dimensions[letter].width = max(width, 2)
    d = os.path.dirname(xlsx_path)
    if d:
        os.makedirs(d, exist_ok=True)
    wb.save(xlsx_path)


def _backup_existing_to_archive(file_path):
    """若 file_path 存在，则复制到同目录下的 99_archive 文件夹，文件名加年月日时分秒后缀。返回备份路径或 None。"""
    if not file_path or not os.path.isfile(file_path):
        return None
    dir_name = os.path.dirname(file_path)
    base_name = os.path.basename(file_path)
    name, ext = os.path.splitext(base_name)
    suffix = datetime.now().strftime("%Y%m%d%H%M%S")
    archive_dir = os.path.join(dir_name, "99_archive")
    os.makedirs(archive_dir, exist_ok=True)
    backup_name = "%s_%s%s" % (name, suffix, ext)
    backup_path = os.path.join(archive_dir, backup_name)
    shutil.copy2(file_path, backup_path)
    return backup_path


def run_t14_3_1_1_1_init(path_t14, path_edc_code):
    """
    生成 t14_3_1-1_1.xlsx。若文件已存在则备份后覆盖。
    path_t14: 输出 xlsx 路径
    path_edc_code: edcdef_code.sas7bdat 或 .xlsx 路径（可为空，则无 AE 码表数据，仅输出空结构或占位行）
    """
    path_t14 = (path_t14 or "").strip()
    if not path_t14:
        raise ValueError("请设置 t14_3_1-1_1.xlsx 输出路径（宏参数 PATH_T14_3_1_1_1 或命令行第 1 个参数）。")

    ae_data = {"aesev_values": [], "aeacn_values": [], "aedis_exist": "N", "aesi_exist": "N"}
    if path_edc_code and os.path.isfile(path_edc_code):
        try:
            ae_data = read_edcdef_code_ae(path_edc_code)
            logger.info("EDCDEF_code AE 解析：aesev_values=%d, aeacn_values=%d, aedis_exist=%s, aesi_exist=%s",
                        len(ae_data["aesev_values"]), len(ae_data["aeacn_values"]),
                        ae_data["aedis_exist"], ae_data["aesi_exist"])
        except Exception as e:
            logger.warning("无法读取 EDCDEF_code，将使用空 AE 数据：%s", e)
    else:
        logger.warning("未提供或找不到 EDCDEF_code 路径，将使用空 AE 数据")

    if os.path.isfile(path_t14):
        backup_path = _backup_existing_to_archive(path_t14)
        if backup_path:
            logger.info("已备份原文件至：%s", backup_path)

    d = os.path.dirname(path_t14)
    if d:
        os.makedirs(d, exist_ok=True)
    rows = build_t14_3_1_1_1_rows(ae_data)
    write_t14_3_1_1_1_xlsx(path_t14, rows)
    logger.info("已初始化 t14_3_1-1_1.xlsx（共 %d 行）：%s", len(rows), path_t14)
    return path_t14


if __name__ == "__main__":
    PATH_T14_3_1_1_1 = r".\t14_3_1-1_1.xlsx"
    PATH_EDCDEF_CODE = r".\edcdef_code.sas7bdat"

    if len(sys.argv) >= 3:
        path_t14 = sys.argv[1]
        path_edc = sys.argv[2]
    else:
        path_t14 = PATH_T14_3_1_1_1
        path_edc = PATH_EDCDEF_CODE or ""
    if not path_t14:
        print("请设置宏参数 PATH_T14_3_1_1_1、PATH_EDCDEF_CODE，或使用：")
        print(" python tfls_metadata_tadae_over.py <t14_3_1-1_1.xlsx路径> <edcdef_code路径>")
        sys.exit(1)
    run_t14_3_1_1_1_init(path_t14, path_edc)
    print("不良事件汇总表 metadata t14_3_1-1_1.xlsx 已生成，文件路径为：", path_t14)
