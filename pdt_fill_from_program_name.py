# -*- coding: utf-8 -*-
"""
基于 program_name.xlsx 与 generate_pdt.sas 的匹配思路，
根据 PDT 中 Title / Output Reference 填写 Program Name 与 SYSPARM Value 列。
"""
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# program_name.xlsx 中使用的 sheet 及对应的 section（与 generate_pdt.sas 一致）
PROGRAM_NAME_SHEETS = [
    "over",
    "s14_1", "s14_2", "s14_3_1", "s14_3_4", "s14_3_5", "s14_4",
    "s16_1_9", "s16_2",
]
# s14_3_2 在 SAS 中由 s14_3_1 派生，此处用同一数据 + section 14.3.2
SECTION_FROM_SHEET = {
    "s14_1": "14.1", "s14_2": "14.2", "s14_3_1": "14.3.1", "s14_3_2": "14.3.2",
    "s14_3_4": "14.3.4", "s14_3_5": "14.3.5", "s14_4": "14.4",
    "s16_1_9": "16.1", "s16_2": "16.2",
}
# 用于匹配的 OUTREF 片段 -> section（取最长匹配）
OUTREF_TO_SECTION = [
    ("14.3.1", "14.3.1"), ("14.3.4", "14.3.4"), ("14.3.2", "14.3.2"), ("14.3.5", "14.3.5"),
    ("14.1", "14.1"), ("14.2", "14.2"), ("14.4", "14.4"),
    ("16.2", "16.2"), ("16.1", "16.1"),
]
# 标题/程序/参数 多值分隔符（与 SAS &esc. 对应，用不常见字符避免与标题内容冲突）
ESC = "\x1e"


def _compress(s):
    """去除空白，用于匹配时与 SAS compress 一致。"""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return re.sub(r"\s+", "", str(s).strip())


def _lowcase(s):
    """转小写，用于匹配。"""
    return (_compress(s)).lower()


def _find_col(df, *candidates):
    """在 DataFrame 列名中查找第一个存在的候选（不区分大小写）。"""
    cols_lower = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = str(cand).strip().lower()
        if key in cols_lower:
            return cols_lower[key]
    return None


def load_program_name_excel(excel_path, lng="cn"):
    """
    读取 program_name.xlsx，构建：
    - over: list of dict {title, pgm, sysparm}
    - sections: dict section -> list of dict {title1, title2, title3, pgm1, pgm2, pgm3, sysparm1, sysparm2, sysparm3}
    lng: 'cn' | 'en'，决定用 title_shell_cn 还是 title_shell_en。
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"program_name.xlsx 不存在: {excel_path}")

    xl = pd.ExcelFile(excel_path)
    over = []
    sections = {}

    title_key = f"title_shell_{lng.lower()}"
    title2_key = f"title2_shell_{lng.lower()}"
    title3_key = f"title3_shell_{lng.lower()}"

    for sheet in PROGRAM_NAME_SHEETS:
        if sheet not in xl.sheet_names:
            continue
        df = pd.read_excel(xl, sheet_name=sheet)
        if df.empty:
            continue

        # 列名不区分大小写查找
        def col(name):
            return _find_col(df, name, name.replace("_", " "))

        if sheet == "over":
            title_col = col("TITLE_SHELL_CN") or col("TITLE_SHELL_EN") or col(title_key)
            pgm_col = col("PGM_SHELL") or col("pgm_shell")
            sys_col = col("SYSPARM_SHELL") or col("sysparm_shell")
            if title_col and pgm_col:
                for _, row in df.iterrows():
                    t = row.get(title_col)
                    p = row.get(pgm_col)
                    sp = row.get(sys_col) if sys_col else ""
                    if pd.notna(t) and str(t).strip():
                        over.append({
                            "title": _compress(t),
                            "pgm": "" if pd.isna(p) else str(p).strip(),
                            "sysparm": "" if pd.isna(sp) else str(sp).strip(),
                        })
            continue

        section = SECTION_FROM_SHEET.get(sheet)
        if not section:
            continue

        c1 = col("title_shell_cn") or col("title_shell_en")
        c2 = col("title2_shell_cn") or col("title2_shell_en")
        c3 = col("title3_shell_cn") or col("title3_shell_en")
        p1 = col("pgm_shell")
        p2 = col("pgm2_shell")
        p3 = col("pgm3_shell")
        s1 = col("sysparm_shell")
        s2 = col("sysparm2_shell")
        s3 = col("sysparm3_shell")

        # 优先使用 lng 指定列
        c1 = col(title_key) or c1
        c2 = col(title2_key) or c2
        c3 = col(title3_key) or c3

        if c1 is None and p1 is None:
            continue

        def _val(r, col_name, default=""):
            if col_name is None:
                return default
            v = r.get(col_name)
            return default if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v).strip()

        rows = []
        for _, row in df.iterrows():
            rows.append({
                "title1": _val(row, c1),
                "title2": _val(row, c2),
                "title3": _val(row, c3),
                "pgm1": _val(row, p1),
                "pgm2": _val(row, p2),
                "pgm3": _val(row, p3),
                "sysparm1": _val(row, s1),
                "sysparm2": _val(row, s2),
                "sysparm3": _val(row, s3),
            })

        if section not in sections:
            sections[section] = []
        sections[section].extend(rows)

    # s14_3_2: 用 s14_3_1 的数据，section 改为 14.3.2，且无 title3/pgm3/sysparm3（SAS 里 drop 了）
    if "14.3.1" in sections and "14.3.2" not in sections:
        s14_3_1_rows = sections["14.3.1"]
        sections["14.3.2"] = [
            {
                "title1": r["title1"],
                "title2": r["title2"],
                "title3": "",
                "pgm1": r["pgm1"],
                "pgm2": r["pgm2"],
                "pgm3": "ladae" if "ladae" in str(r.get("pgm2", "")).lower() else "",
                "sysparm1": r["sysparm1"],
                "sysparm2": r["sysparm2"],
                "sysparm3": "",
            }
            for r in s14_3_1_rows
        ]

    return {"over": over, "sections": sections}


def _get_section_from_outref(outref):
    """从 Output Reference 字符串解析出 section（14.1, 14.3.1, 16.2 等）。"""
    s = _compress(outref)
    for ref_part, section in OUTREF_TO_SECTION:
        if ref_part in s or section in s:
            return section
    return None


def _match_single_level(outtitle_norm, shell_rows, outtype, outpop):
    """单层匹配：14.1 / 16.2。返回 (pgmnamdv, outsysp, len_remain) 或 None。"""
    best = None
    for r in shell_rows:
        t1 = _lowcase(r["title1"])
        if not t1:
            continue
        if t1 not in outtitle_norm:
            continue
        remain = outtitle_norm.replace(t1, "", 1)
        len_remain = len(_compress(remain))
        if best is None or len_remain < best[2]:
            best = (r["pgm1"], r["sysparm1"] or "", len_remain)
    return best


def _match_three_level(outtitle_norm, outtitle_orig, outref, outtype, outpop, shell_rows):
    """三层匹配：14.3.1 / 14.3.4。返回 (pgmnamdv, outsysp, len_remain) 或 None。"""
    best = None
    for r in shell_rows:
        t1 = _lowcase(r["title1"])
        t2 = _lowcase(r["title2"])
        t3 = _lowcase(r["title3"])
        if not t1 or not t2 or not t3:
            continue
        if t1 not in outtitle_norm or t2 not in outtitle_norm or t3 not in outtitle_norm:
            continue
        remain = outtitle_norm
        for t in (t1, t2, t3):
            remain = remain.replace(t, "", 1)
        len_remain = len(_compress(remain))

        sysparm1 = (r["sysparm1"] or "").strip()
        sysparm2 = (r["sysparm2"] or "").strip()
        sysparm3 = (r["sysparm3"] or "").strip()

        if "发生率>=10" in outtitle_orig:
            sysparm3 = sysparm3.replace("@sevcol", "@pctlmt=10@sevcol")
        elif "发生率>=5" in outtitle_orig:
            sysparm3 = sysparm3.replace("@sevcol", "@pctlmt=5@sevcol")
        else:
            sysparm3 = sysparm3.replace("@sevcol", "@pctlmt=0@sevcol")

        if outtype == "Table":
            sysparm2 = sysparm2.replace("lbcat", "parcat1")
            if (r.get("pgm3") or "").strip() == "byvis":
                sysparm2 = sysparm2 + ' and  anl01fl="Y"'
            elif (r.get("pgm3") or "").strip() == "shift_byvis":
                sysparm2 = sysparm2 + ' and  anl03fl="Y"'
            elif (r.get("pgm3") or "").strip() == "shift":
                sysparm2 = sysparm2 + ' and  anl02fl="Y"'
            if "Safety Set" in (outpop or "") or "安全性分析集" in (outpop or ""):
                if "14.3.1" in (outref or ""):
                    outsysp = "@".join(filter(None, [sysparm1, sysparm2, 'adsl_scr=%str(1)', sysparm3]))
                elif "14.3.4" in (outref or ""):
                    sysparm2_quoted = f'datascr=%quote({sysparm2})'
                    outsysp = "@".join(filter(None, [sysparm1, sysparm2_quoted, 'adslscr=%quote(saffl="Y")', sysparm3]))
                else:
                    outsysp = "@".join(filter(None, [sysparm1, sysparm2, sysparm3]))
            else:
                outsysp = "@".join(filter(None, [sysparm1, sysparm2, sysparm3]))
        else:
            outsysp = " ".join(filter(None, [sysparm1, sysparm2, sysparm3]))

        pgm = "_".join(filter(None, [r["pgm1"], r["pgm2"], r["pgm3"]]))
        if best is None or len_remain < best[2]:
            best = (pgm, outsysp, len_remain)
    return best


def _match_two_level(outtitle_norm, outtitle_orig, outref, outtype, outpop, shell_rows):
    """两层匹配：14.3.2 / 14.3.5。返回 (pgmnamdv, outsysp, len_remain) 或 None。"""
    best = None
    for r in shell_rows:
        t1 = _lowcase(r["title1"])
        t2 = _lowcase(r["title2"])
        if not t1 or not t2:
            continue
        if t1 not in outtitle_norm or t2 not in outtitle_norm:
            continue
        remain = outtitle_norm.replace(t1, "", 1).replace(t2, "", 1)
        len_remain = len(_compress(remain))

        sysparm1 = (r["sysparm1"] or "").strip()
        sysparm2 = (r["sysparm2"] or "").strip()
        if "14.3.2" in (outref or ""):
            sysparm2 = re.sub(r"data_scr=%str\s*", "", sysparm2)

        if outtype == "Table":
            sysparm1 = (sysparm1
                .replace("vscat", "parcat1").replace("pecat", "parcat1")
                .replace("egscat", "parcat2").replace("pdcat", "parcat1"))
            if (r.get("pgm2") or "").strip() == "byvis":
                sysparm1 = sysparm1 + ' and  anl01fl="Y"'
            elif (r.get("pgm2") or "").strip() == "shift_byvis":
                sysparm1 = sysparm1 + ' and  anl03fl="Y"'
            elif (r.get("pgm2") or "").strip() == "shift":
                sysparm1 = sysparm1 + ' and  anl02fl="Y"'
            if "14.3.5" in (outref or ""):
                sysparm1_quoted = f'datascr=%quote({sysparm1})'
                outsysp = "@".join(filter(None, [sysparm1_quoted, 'adslscr=%quote(saffl="Y")', sysparm2]))
            else:
                outsysp = "@".join(filter(None, [sysparm1, sysparm2]))
        else:
            outsysp = " ".join(filter(None, [sysparm1, sysparm2]))

        pgm = "_".join(filter(None, [r["pgm1"], r["pgm2"]]))
        if best is None or len_remain < best[2]:
            best = (pgm, outsysp, len_remain)
    return best


def match_pdt_row(outref, outtitle, outtype, outpop, program_data, lng="cn"):
    """
    对单条 PDT 行根据 Output Reference、Title、Output Type、Population 匹配程序名与 SYSPARM。
    返回 (program_name, sysparm_value)，未匹配时为 ("", "")。
    """
    sections = program_data["sections"]
    over_list = program_data["over"]

    outref = (outref or "").strip()
    outtitle_orig = (outtitle or "").strip()
    outtitle_norm = _lowcase(outtitle_orig)
    outtype = (outtype or "").strip()
    outpop = (outpop or "").strip()

    section = _get_section_from_outref(outref)
    if not section or section not in sections:
        return "", ""

    shell_rows = sections[section]
    if not shell_rows:
        return "", ""

    best = None
    if section in ("14.1", "16.2"):
        best = _match_single_level(outtitle_norm, shell_rows, outtype, outpop)
    elif section in ("14.3.1", "14.3.4"):
        best = _match_three_level(outtitle_norm, outtitle_orig, outref, outtype, outpop, shell_rows)
    elif section in ("14.3.2", "14.3.5"):
        best = _match_two_level(outtitle_norm, outtitle_orig, outref, outtype, outpop, shell_rows)
    else:
        # 14.2, 14.4, 16.1 等：按单层处理
        best = _match_single_level(outtitle_norm, shell_rows, outtype, outpop)

    if best is None:
        return "", ""

    pgmnamdv, outsysp = best[0], best[1]
    if not pgmnamdv:
        return "", ""

    # Over 叠加：若 Title 包含 over 的 title，则程序名后追加 _over_pgm，并合并 over_sys
    for o in over_list:
        ot = o["title"]
        if not ot:
            continue
        if _compress(outtitle_orig).find(ot) >= 0 or outtitle_norm.find(_lowcase(ot)) >= 0:
            if o.get("pgm"):
                pgmnamdv = f"{pgmnamdv}_{o['pgm']}"
            if o.get("sysparm"):
                outsysp = outsysp.replace('adsl_scr=%str(1)', f'adsl_scr=%str({o["sysparm"]})')
                outsysp = outsysp.replace('saffl="Y"', f'saffl="Y" and {o["sysparm"]}')
            break

    if pgmnamdv and not pgmnamdv.endswith(".sas"):
        pgmnamdv = pgmnamdv + ".sas"
    return pgmnamdv, outsysp


def _normalize_header(val):
    if val is None:
        return ""
    s = str(val).strip().replace("\u200b", "").replace("\ufeff", "")
    return " ".join(s.split())


def fill_pdt_program_and_sysparm(
    pdt_path,
    program_name_path,
    lng="cn",
    sheet_name="Deliverables",
    program_name_col="Program Name",
    sysparm_col="SYSPARM Value",
    backup=True,
):
    """
    读取 PDT 的 Deliverables sheet，对 Category=Output 的每一行，
    根据 Title / Output Reference 用 program_name.xlsx 匹配并填写 Program Name、SYSPARM Value。

    Args:
        pdt_path: PDT 文件路径（.xlsx）
        program_name_path: program_name.xlsx 路径
        lng: 'cn' | 'en'，与 program_name 中标题列一致
        sheet_name: PDT 中表名，默认 Deliverables
        program_name_col: 程序名列的显示名，默认 "Program Name"
        sysparm_col: SYSPARM 列的显示名，默认 "SYSPARM Value"
        backup: 是否在写回前备份到 99_archive

    Returns:
        (success: bool, message: str)
    """
    try:
        program_data = load_program_name_excel(program_name_path, lng=lng)
    except Exception as e:
        return False, f"读取 program_name.xlsx 失败: {e}"

    pdt_path = Path(pdt_path)
    if not pdt_path.exists():
        return False, f"PDT 文件不存在: {pdt_path}"

    if backup:
        try:
            from tfls_pdt_gen import _backup_pdt
            _backup_pdt(str(pdt_path))
        except Exception:
            pass

    wb = load_workbook(pdt_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return False, f"PDT 中未找到 sheet: {sheet_name}"

    ws = wb[sheet_name]
    header_row, col_name_to_idx = _find_header_row_and_cols(ws, program_name_col, sysparm_col)
    if header_row is None or "Category" not in col_name_to_idx:
        wb.close()
        return False, "Deliverables 中未找到表头（Category 等）"

    cat_col = col_name_to_idx.get("Category")
    outref_col = col_name_to_idx.get("Output Reference")
    title_col = col_name_to_idx.get("Title")
    outtype_col = col_name_to_idx.get("Output Type")
    outpop_col = col_name_to_idx.get("Population")
    pgm_col = col_name_to_idx.get(program_name_col) or col_name_to_idx.get("Program Name")
    sysparm_out_col = col_name_to_idx.get(sysparm_col) or col_name_to_idx.get("SYSPARM Value")

    if not outref_col or not title_col:
        wb.close()
        return False, "未找到 Output Reference 或 Title 列"
    if not pgm_col or not sysparm_out_col:
        wb.close()
        return False, "未找到 Program Name 或 SYSPARM Value 列，请确认 PDT 表头包含这两列"

    filled = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cat_val = ws.cell(row=row_idx, column=cat_col).value
        if cat_val is None or _normalize_header(cat_val) != "Output":
            continue
        outref = ws.cell(row=row_idx, column=outref_col).value
        outtitle = ws.cell(row=row_idx, column=title_col).value
        outtype = ws.cell(row=row_idx, column=outtype_col).value if outtype_col else ""
        outpop = ws.cell(row=row_idx, column=outpop_col).value if outpop_col else ""
        pgmnamdv, outsysp = match_pdt_row(outref, outtitle, outtype, outpop, program_data, lng=lng)
        if pgmnamdv or outsysp:
            ws.cell(row=row_idx, column=pgm_col, value=pgmnamdv)
            ws.cell(row=row_idx, column=sysparm_out_col, value=outsysp)
            filled += 1

    wb.save(pdt_path)
    wb.close()
    return True, f"已根据 Title 填写 Program Name 与 SYSPARM Value，共处理 {filled} 行 Output。"


def _find_header_row_and_cols(ws, program_name_col="Program Name", sysparm_col="SYSPARM Value"):
    """查找表头行及列索引，包含 Program Name / SYSPARM Value。"""
    from tfls_pdt_gen import PDT_DELIVERABLES_COLS, PDT_COL_ALIASES, _find_header_row_and_cols as _base
    header_row, col_name_to_idx = _base(ws)
    if header_row is None:
        return None, {}
    # 同一行补充 Program Name / SYSPARM Value 列（PDT 可能有这两列但不在 PDT_DELIVERABLES_COLS 中）
    for col_idx, cell in enumerate(ws[header_row], start=1):
        val = cell.value
        if val is None:
            continue
        norm = _normalize_header(val)
        if norm in ("Program Name", "PGMNAMDV"):
            col_name_to_idx["Program Name"] = col_idx
        elif norm in ("SYSPARM Value", "OUTSYSP"):
            col_name_to_idx["SYSPARM Value"] = col_idx
    return header_row, col_name_to_idx


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("用法: python pdt_fill_from_program_name.py <PDT.xlsx> <program_name.xlsx> [lng=cn]")
        sys.exit(1)
    pdt = sys.argv[1]
    pnm = sys.argv[2]
    lng = sys.argv[3] if len(sys.argv) > 3 else "cn"
    ok, msg = fill_pdt_program_and_sysparm(pdt, pnm, lng=lng)
    print(msg)
    sys.exit(0 if ok else 1)
