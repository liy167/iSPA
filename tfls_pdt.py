# -*- coding: utf-8 -*-
"""
TFLs 页面 - 生成PDT 弹窗逻辑（独立模块，避免 SASEG_GUI 过于庞大）
"""
import os
import re
import shutil
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def convert_windows_path_to_linux(win_path):
    """将 Windows 路径（Z:\\...）转为服务器 Linux 路径。"""
    if not win_path:
        return win_path
    linux_base = '/u01/app/sas/sas9.4/DocumentRepository/DDT'
    s = win_path.strip()
    if s.upper().startswith('Z:\\') or s.upper().startswith('Z:/'):
        s = linux_base + s[2:]
    s = s.replace('\\', '/')
    return s


# ---------- TOC 生成（初版TOC）所需常量与函数 ----------
# 当问题1 仅选择一个答案时，TOC.xlsx 命名规则（与 OUTTITLE 一致，不追加设计类型后缀）：
#   OUTTITLE：取模板标题（Title_CN/Title_EN），不追加设计类型后缀；[Analyte]/[AEACN] 按规则替换为具体值。
#   OUTREF：不追加设计类型序号，为 Template# 或 Template#.aeacn序号 或 Template#.analyte序号。
#           例：14.1、14.3.1-5.1（AEACN 第1个）、14.1.2（Analyte 第2个）。多选设计类型时才追加 .设计类型序号。
_TOC_COLS = [
    "Template#", "Output Type", "Title_CN", "Title_EN", "Population",
    "Footnotes_CN", "Footnotes_EN", "Category_CN", "SAD", "FE", "MAD", "BE", "MB"
]
_TOC_SHEET_COLS = ["OUTTYPE", "OUTREF", "OUTTITLE", "OUTPOP", "OUTNOTE"]
_EXCLUDED_CATEGORY_CN = {"QT分析", "C-QT分析", "PK浓度", "PK参数", "PD分析", "ADA分析"}
_ENDPOINT_TO_CATEGORY = {
    "PK浓度(血)": "PK浓度", "PK浓度(尿)": "PK浓度", "PK浓度(粪)": "PK浓度",
    "PK参数(血)": "PK参数", "PK参数(尿)": "PK参数", "PK参数(粪)": "PK参数",
    "PD分析": "PD分析", "ADA分析": "ADA分析", "QT分析": "QT分析",
}
_DESIGN_TYPE_COLS = ["SAD", "FE", "MAD", "BE", "MB"]
_PLACEHOLDER_ANALYTE = "[Analyte]"
_PLACEHOLDER_AEACN = "[AEACN]"
_AEACN_EXCLUDED_LABELS = frozenset(
    s.strip().upper() for s in
    ("剂量不变", "不适用", "DOSE NOT CHANGED", "NOT APPLICABLE")
)
_SUBTYPE_TERMS = {
    "血": ["血", "Blood", "blood", "血浆", "Plasma", "plasma"],
    "尿": ["尿", "Urine", "urine"],
    "粪": ["粪", "粪便", "Feces", "feces", "Stool", "stool"],
}


def _toc_read_lng(setup_path):
    """从 setup.xlsx 的 Macro Variables sheet 读取 LNG。"""
    wb = load_workbook(setup_path, read_only=True, data_only=True)
    if "Macro Variables" not in wb.sheetnames:
        wb.close()
        return ""
    ws = wb["Macro Variables"]
    lng_val = ""
    for row in ws.iter_rows(min_row=1, max_col=3):
        b_val = row[1].value if len(row) > 1 else None
        if b_val is not None and str(b_val).strip().upper() == "LNG":
            c_val = row[2].value if len(row) > 2 else None
            lng_val = str(c_val).strip() if c_val is not None else ""
            break
    wb.close()
    return lng_val


def _toc_is_chinese_lng(lng_val):
    """根据 LNG 值判断是否使用中文列。"""
    if not lng_val:
        return True
    v = lng_val.upper()
    if v in ("CHN", "CN", "CHINESE", "中文", "ZH", "ZH-CN"):
        return True
    return False


def _toc_normalize_header(h):
    if h is None:
        return ""
    s = str(h).strip().replace("\u200b", "").replace("\ufeff", "")
    return " ".join(s.split())


def _toc_normalize_analyte_placeholder(s):
    if not s:
        return s
    s = str(s).replace("<Analyte分析物>", _PLACEHOLDER_ANALYTE).replace("<Analyte>", _PLACEHOLDER_ANALYTE)
    return s


def _edcdef_code_aeacn_labels(edcdef_code_path):
    """
    从 EDCDEF_code.sas7bdat（或 .xlsx）中读取 CODE_NAME='AEACN' 的 CODE_LABEL 列表。
    筛选条件：CODE_NAME='AEACN' 且 CODE_LABEL 不在 ('剂量不变','不适用','DOSE NOT CHANGED','NOT APPLICABLE')；
    按 CODE_ORDER 排序后返回 CODE_LABEL 值列表。文件不存在或读取失败返回 []。
    """
    if not edcdef_code_path or not os.path.isfile(edcdef_code_path):
        return []
    ext = os.path.splitext(edcdef_code_path)[1].lower()
    try:
        if ext == ".sas7bdat":
            import pyreadstat
            df, _ = pyreadstat.read_sas7bdat(edcdef_code_path)
        elif ext in (".xlsx", ".xls"):
            import pandas as pd
            df = pd.read_excel(edcdef_code_path, header=0)
        else:
            return []
    except Exception:
        return []
    if df is None or df.empty:
        return []
    cols_upper = {str(c).upper(): c for c in df.columns}
    code_name_col = cols_upper.get("CODE_NAME") or cols_upper.get("CODE_NAME_CHN")
    code_label_col = cols_upper.get("CODE_LABEL")
    code_order_col = cols_upper.get("CODE_ORDER") or cols_upper.get("CODE_ORDER_R")
    if code_name_col is None or code_label_col is None:
        return []
    name_vals = df[code_name_col].astype(str).str.strip()
    aeacn_mask = name_vals.str.upper() == "AEACN"
    if not aeacn_mask.any():
        return []
    sub = df.loc[aeacn_mask].copy()
    label_vals = sub[code_label_col].astype(str).str.strip()
    excluded = _AEACN_EXCLUDED_LABELS
    keep = ~label_vals.str.upper().isin(excluded)
    sub = sub.loc[keep]
    if sub.empty:
        return []
    if code_order_col is not None:
        try:
            sub = sub.sort_values(by=code_order_col)
        except Exception:
            pass
    return sub[code_label_col].astype(str).str.strip().tolist()


def _edcdef_ecrf_has_ae_aedis(edcdef_ecrf_path):
    """
    读取 utility\\metadata\\EDCDEF_ecrf.sas7bdat，当 EDC_DATA='AE' 时是否存在 EDC_VARIABLE='AEDIS'。
    若文件不存在或读取失败返回 False。
    """
    if not edcdef_ecrf_path or not os.path.isfile(edcdef_ecrf_path):
        return False
    try:
        import pyreadstat
        df, _ = pyreadstat.read_sas7bdat(edcdef_ecrf_path)
    except Exception:
        return False
    if df is None or df.empty:
        return False
    # 列名可能为大写或混合
    cols = {c.upper(): c for c in df.columns}
    edc_data_col = cols.get("EDC_DATA") or cols.get("EDC_DATA".lower())
    edc_var_col = cols.get("EDC_VARIABLE") or cols.get("EDC_VARIABLE".lower())
    if edc_data_col is None or edc_var_col is None:
        return False
    ae_mask = df[edc_data_col].astype(str).str.strip().str.upper() == "AE"
    if not ae_mask.any():
        return False
    ae_df = df.loc[ae_mask]
    aedis_mask = ae_df[edc_var_col].astype(str).str.strip().str.upper() == "AEDIS"
    return aedis_mask.any()


def _toc_read_rows(toc_path):
    """读取 TOC 的 PH1 sheet，返回行列表（每行为 dict）。"""
    wb = load_workbook(toc_path, read_only=True, data_only=True)
    if "PH1" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["PH1"]
    col_idx = {}
    alias_map = {"Footnote_CN": "Footnotes_CN", "Footnote_EN": "Footnotes_EN"}
    rows = []
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if not col_idx:
            for i, h in enumerate(vals):
                norm = _toc_normalize_header(h)
                if norm in _TOC_COLS:
                    col_idx[norm] = i
                elif norm in alias_map:
                    col_idx[alias_map[norm]] = i
            continue
        row_dict = {col_name: vals[idx] if idx < len(vals) else None for col_name, idx in col_idx.items()}
        rows.append(row_dict)
    wb.close()
    return rows


def _toc_filter_and_expand_rows(toc_rows, design_types, endpoints, use_cn, analyte_names=None, aeacn_labels=None):
    """按设计类型、终点、分析物、[AEACN] 筛选并展开，生成 TOC 行。aeacn_labels 来自 EDCDEF_code CODE_NAME='AEACN' 的 CODE_LABEL 列表（已排序）。"""
    use_title = "Title_CN" if use_cn else "Title_EN"
    use_footnotes = "Footnotes_CN" if use_cn else "Footnotes_EN"
    placeholder = _PLACEHOLDER_ANALYTE
    analytes = [a.strip() for a in (analyte_names or "").split("|") if a.strip()] if analyte_names else []
    aeacn_list = list(aeacn_labels) if aeacn_labels else []
    base_categories = set()
    for ep in endpoints:
        if ep in _ENDPOINT_TO_CATEGORY:
            base_categories.add(_ENDPOINT_TO_CATEGORY[ep])
    selected_subtypes = set()
    for ep in ["PK浓度(血)", "PK浓度(尿)", "PK浓度(粪)", "PK参数(血)", "PK参数(尿)", "PK参数(粪)"]:
        if ep in endpoints:
            selected_subtypes.add(ep[-2])
    excluded_subtype_terms = []
    for st, terms in _SUBTYPE_TERMS.items():
        if st not in selected_subtypes:
            excluded_subtype_terms.extend(terms)

    def _title_contains_excluded(row):
        if not excluded_subtype_terms or not selected_subtypes:
            return False
        title_cn = str(row.get("Title_CN") or "")
        title_en = str(row.get("Title_EN") or "")
        text = title_cn + " " + title_en
        text_lower = text.lower()
        for term in excluded_subtype_terms:
            if len(term) <= 2:
                if term in title_cn or term in title_en:
                    return True
            elif term.lower() in text_lower:
                return True
        return False

    selected_rows = []
    for r in toc_rows:
        cat_cn = (r.get("Category_CN") or "").strip()
        if not cat_cn:
            continue
        if cat_cn in _EXCLUDED_CATEGORY_CN and cat_cn not in base_categories:
            continue
        if cat_cn in ("PK浓度", "PK参数") and _title_contains_excluded(r):
            continue
        selected_rows.append(r)

    result = []
    for r in selected_rows:
        template_num = str(r.get("Template#") or "").strip()
        output_type = r.get("Output Type") or ""
        title = _toc_normalize_analyte_placeholder(r.get(use_title) or "")
        population = r.get("Population") or ""
        footnotes = r.get(use_footnotes) or ""
        dt_cols_present = [c for c in _DESIGN_TYPE_COLS if c in r and r.get(c) is not None]
        for dt in _DESIGN_TYPE_COLS:
            if dt not in design_types:
                continue
            dt_val = r.get(dt)
            if dt_cols_present and (dt_val is None or (isinstance(dt_val, str) and not str(dt_val).strip())):
                continue
            design_ordinal = design_types.index(dt) + 1
            single_design = len(design_types) == 1
            base_title = title if single_design else (f"{title} - {dt}" if title else dt)

            def _build_out_ref(analyte_idx=None, aeacn_idx=None):
                parts = [template_num] if template_num else []
                if aeacn_idx is not None:
                    parts.append(str(aeacn_idx))
                if analyte_idx is not None:
                    parts.append(str(analyte_idx))
                if not single_design:
                    parts.append(str(design_ordinal))
                return ".".join(parts)

            if _PLACEHOLDER_AEACN in base_title and aeacn_list:
                for idx, label in enumerate(aeacn_list, start=1):
                    title_with_dt = base_title.replace(_PLACEHOLDER_AEACN, label)
                    out_ref = _build_out_ref(aeacn_idx=idx)
                    result.append({
                        "Output Type": output_type, "Title": title_with_dt, "Population": population,
                        "Footnotes": footnotes, "Output Reference": out_ref,
                    })
            elif _PLACEHOLDER_AEACN in base_title and not aeacn_list:
                continue
            elif placeholder in base_title and analytes:
                for idx, analyte in enumerate(analytes, start=1):
                    title_with_dt = base_title.replace(placeholder, analyte)
                    out_ref = _build_out_ref(analyte_idx=idx)
                    result.append({
                        "Output Type": output_type, "Title": title_with_dt, "Population": population,
                        "Footnotes": footnotes, "Output Reference": out_ref,
                    })
            else:
                title_with_dt = base_title.replace(_PLACEHOLDER_AEACN, "").replace(placeholder, analytes[0] if analytes else "") if (_PLACEHOLDER_AEACN in base_title or placeholder in base_title) else base_title
                out_ref = _build_out_ref()
                result.append({
                    "Output Type": output_type, "Title": title_with_dt, "Population": population,
                    "Footnotes": footnotes, "Output Reference": out_ref,
                })
    return result


def gen_toc_study(template_path, study_path, setup_path, design_types, endpoints, analyte_names=None, edcdef_ecrf_path=None, edcdef_code_path=None):
    """
    根据 TOC_template.xlsx 与前三个问题（设计类型、终点、分析物）、[AEACN]，筛选并展开后生成 TOC.xlsx。
    TOC sheet 列：OUTTYPE, OUTREF, OUTTITLE, OUTPOP, OUTNOTE。
    edcdef_ecrf_path: 若提供，则根据 EDCDEF_ecrf.sas7bdat 中 EDC_DATA='AE' 时是否存在 EDC_VARIABLE='AEDIS'，
                      决定是否保留 Template# 为 14.3.1-5.1 / 14.3.1-5.2 的行；不存在则不保留。
    edcdef_code_path: 若提供，则从中读取 CODE_NAME='AEACN' 的 CODE_LABEL 列表，用于展开 [AEACN] 占位符行（Template# 加后缀 .1/.2/...）。
    """
    use_cn = True
    if setup_path and os.path.isfile(setup_path):
        lng_val = _toc_read_lng(setup_path)
        use_cn = _toc_is_chinese_lng(lng_val)
    toc_rows = _toc_read_rows(template_path)
    if not toc_rows:
        return False, "TOC_template 的 PH1 sheet 未找到或为空"
    if edcdef_ecrf_path:
        keep_14_3_1_5 = _edcdef_ecrf_has_ae_aedis(edcdef_ecrf_path)
        if not keep_14_3_1_5:
            toc_rows = [
                r for r in toc_rows
                if str(r.get("Template#") or "").strip() not in ("14.3.1-5.1", "14.3.1-5.2")
            ]
    aeacn_labels = _edcdef_code_aeacn_labels(edcdef_code_path) if edcdef_code_path else None
    new_rows = _toc_filter_and_expand_rows(toc_rows, design_types, endpoints, use_cn, analyte_names, aeacn_labels=aeacn_labels)
    toc_sheet_rows = [
        {"OUTTYPE": r.get("Output Type") or "", "OUTREF": r.get("Output Reference") or "",
         "OUTTITLE": r.get("Title") or "", "OUTPOP": r.get("Population") or "", "OUTNOTE": r.get("Footnotes") or ""}
        for r in new_rows
    ]
    d = os.path.dirname(study_path)
    if d:
        os.makedirs(d, exist_ok=True)
    if os.path.isfile(study_path):
        study_dir = os.path.dirname(os.path.abspath(study_path))
        archive_dir = os.path.join(study_dir, "99_archive")
        os.makedirs(archive_dir, exist_ok=True)
        base_name = os.path.splitext(os.path.basename(study_path))[0]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(study_path, os.path.join(archive_dir, f"{base_name}_{ts}.xlsx"))
    wb = Workbook()
    ws = wb.active
    ws.title = "TOC"
    for col_idx, col_name in enumerate(_TOC_SHEET_COLS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(toc_sheet_rows, start=2):
        for col_idx, col_name in enumerate(_TOC_SHEET_COLS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
    def _cell_width(val):
        if val is None:
            return 0
        s = str(val)
        return sum(2 if "\u4e00" <= c <= "\u9fff" else 1 for c in s)
    for col_idx in range(1, len(_TOC_SHEET_COLS) + 1):
        col_letter = get_column_letter(col_idx)
        max_w = max((_cell_width(ws.cell(row=row_idx, column=col_idx).value) for row_idx in range(1, ws.max_row + 1)), default=0)
        ws.column_dimensions[col_letter].width = min(55, max(8, max_w + 2))
    wb.save(study_path)
    wb.close()
    return True, "已生成 TOC.xlsx（TOC sheet 共 %d 行）。" % len(toc_sheet_rows)


def show_pdt_dialog(gui):
    """
    显示「生成PDT」弹窗（仅保留原第三步：基于 TOC.xlsx 与项目层面 PDT.xlsx，初版PDT/编辑）。
    gui: 主窗口实例，需有 .root, .get_current_path(), .selected_paths, .update_status()
    """
    dlg = tk.Toplevel(gui.root)
    dlg.title("生成PDT")
    dlg.geometry("1000x200")
    dlg.resizable(True, True)
    dlg.transient(gui.root)
    dlg.grab_set()
    dlg.configure(bg="#f0f0f0")

    main = tk.Frame(dlg, padx=20, pady=16, bg="#f0f0f0")
    main.pack(fill=tk.BOTH, expand=True)

    base_path = gui.get_current_path()
    p3, p4 = (gui.selected_paths[2] or ""), (gui.selected_paths[3] or "")
    default_pdt = os.path.join(base_path, "utility", "documentation", f"{p3}_{p4}_PDT.xlsx" if (p3 or p4) else "项目层面_PDT.xlsx")
    default_toc_study = os.path.join(base_path, "utility", "documentation", "03_statistics", "TOC.xlsx")

    def run_gen(toc_widget, pdt_widget):
        """点击初版PDT：通过 linux_sas_call_from_python 在 Linux 服务器上执行 25_generate_pdt_call.sas（路径 = 前四个下拉框 + utility/tools/25_generate_pdt_call.sas）。"""
        base_4 = getattr(gui, "z_drive", "Z:\\")
        for i in range(4):
            if getattr(gui, "selected_paths", None) and i < len(gui.selected_paths) and gui.selected_paths[i]:
                base_4 = os.path.join(base_4, gui.selected_paths[i])
        sas_script_win = os.path.join(base_4, "utility", "tools", "25_generate_pdt_call.sas")
        linux_path = convert_windows_path_to_linux(sas_script_win)

        try:
            from linux_sas_call_from_python import run_sas
        except ImportError as e:
            messagebox.showerror("错误", "无法导入 linux_sas_call_from_python（请确保该模块在项目目录下且已安装 saspy）。\n\n%s" % e)
            return

        try:
            has_issue = run_sas(sas_script_win, check_log=True)
        except Exception as e:
            messagebox.showerror("错误", "调用 SAS 程序时出错：%s" % e)
            return

        # 日志审阅（ERROR/WARNING 弹窗、是否打开日志）已由 linux_sas_call_from_python 完成，此处仅更新 GUI 状态栏
        gui.update_status("25_generate_pdt_call.sas 已执行完成（有 ERROR/WARNING 时已由日志审阅窗口提示）。" if has_issue else "已在 Linux 服务器执行 25_generate_pdt_call.sas。")

    def on_open_edit(pdt_widget):
        p = pdt_widget.get().strip()
        if p and os.path.isfile(p):
            try:
                os.startfile(p)
                gui.update_status("已打开: " + os.path.basename(p))
            except Exception as e:
                messagebox.showerror("错误", "无法打开文件: %s" % e)
        else:
            messagebox.showwarning("提示", "请先选择有效的项目层面PDT.xlsx 路径，或先点击「初版PDT」生成后再打开。")

    # ---------- 基于 TOC 生成 PDT（唯一一步） ----------
    step3_title = tk.Label(main, text="基于TOC.xlsx，并在原项目层面PDT基础上添加TFLs行，生成最新版PDT。", font=("Microsoft YaHei UI", 10, "bold"), fg="#333333", bg="#f0f0f0", wraplength=1240)
    step3_title.pack(anchor="w", pady=(0, 10))

    row_toc_s3 = tk.Frame(main, bg="#f0f0f0")
    row_toc_s3.pack(anchor="w", fill=tk.X, pady=(0, 6))
    tk.Label(row_toc_s3, text="TOC.xlsx：", font=("Microsoft YaHei UI", 9), width=22, anchor="w", bg="#f0f0f0").pack(side=tk.LEFT, padx=(0, 4))
    toc_entry_s3 = tk.Entry(row_toc_s3, width=72, font=("Microsoft YaHei UI", 9))
    toc_entry_s3.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
    toc_entry_s3.insert(0, default_toc_study)

    def browse_toc_s3():
        path = filedialog.askopenfilename(title="选择 TOC.xlsx", filetypes=[("Excel", "*.xlsx"), ("All", "*.*")], initialdir=os.path.dirname(default_toc_study) or base_path)
        if path:
            toc_entry_s3.delete(0, tk.END)
            toc_entry_s3.insert(0, path)

    tk.Button(row_toc_s3, text="浏览...", command=browse_toc_s3, width=8, font=("Microsoft YaHei UI", 9)).pack(side=tk.LEFT)

    row_pdt_s3 = tk.Frame(main, bg="#f0f0f0")
    row_pdt_s3.pack(anchor="w", fill=tk.X, pady=(0, 14))
    tk.Label(row_pdt_s3, text="项目层面PDT.xlsx：", font=("Microsoft YaHei UI", 9), width=22, anchor="w", bg="#f0f0f0").pack(side=tk.LEFT, padx=(0, 4))
    pdt_entry_s3 = tk.Entry(row_pdt_s3, width=72, font=("Microsoft YaHei UI", 9))
    pdt_entry_s3.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
    pdt_entry_s3.insert(0, default_pdt)

    def browse_pdt_s3():
        path = filedialog.askopenfilename(title="选择 项目层面PDT.xlsx", filetypes=[("Excel", "*.xlsx"), ("All", "*.*")], initialdir=os.path.dirname(default_pdt) or base_path)
        if path:
            pdt_entry_s3.delete(0, tk.END)
            pdt_entry_s3.insert(0, path)

    tk.Button(row_pdt_s3, text="浏览...", command=browse_pdt_s3, width=8, font=("Microsoft YaHei UI", 9)).pack(side=tk.LEFT)

    btn_frame_s3 = tk.Frame(main, bg="#f0f0f0")
    btn_frame_s3.pack(anchor="w", pady=(8, 0))

    tk.Button(btn_frame_s3, text="初版PDT", command=lambda: run_gen(toc_entry_s3, pdt_entry_s3), width=10, font=("Microsoft YaHei UI", 9)).pack(side=tk.LEFT, padx=(0, 8))
    tk.Button(btn_frame_s3, text="编辑", command=lambda: on_open_edit(pdt_entry_s3), width=10, font=("Microsoft YaHei UI", 9)).pack(side=tk.LEFT)
