# -*- coding: utf-8 -*-
"""
PDT 生成模块：备份原PDT，按 TOC 与用户选择生成 Deliverables sheet 中 Category=Output 的行。
"""
import os
import shutil
from datetime import datetime
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, get_column_letter
from openpyxl.styles import PatternFill

# 列名常量
TOC_COLS = [
    "Template#", "Output Type", "Title_CN", "Title_EN", "Population",
    "Footnotes_CN", "Footnotes_EN", "Category_CN", "SAD", "FE", "MAD", "BE", "MB"
]
PDT_DELIVERABLES_COLS = ["Category", "Output Type", "Title", "Population", "Footnotes", "Output Reference"]
# PDT 列别名（实际文件可能使用不同列名）
PDT_COL_ALIASES = {
    "OUTCAT": "Category",
    "OUTTYPE": "Output Type",
    "OUTREF": "Output Reference",
    "OUTTITLE": "Title",
    "OUTPOP": "Population",
    "OUTFNOTE": "Footnotes",
    "PGMLEVEL": "Validation Level",
    "USERDEV": "Developers",
    "USERQC": "Validators",
    "OUTSTS": "Output Status",
    "STASCHK": "Validated by Programmer/Statistician",
}
EXCLUDED_CATEGORY_CN = {"QT分析", "C-QT分析", "PK浓度", "PK参数", "PD分析", "ADA分析"}

# 终点 -> Category_CN 映射（用于额外添加）
ENDPOINT_TO_CATEGORY = {
    "PK浓度(血)": "PK浓度",
    "PK浓度(尿)": "PK浓度",
    "PK浓度(粪)": "PK浓度",
    "PK参数(血)": "PK参数",
    "PK参数(尿)": "PK参数",
    "PK参数(粪)": "PK参数",
    "PD分析": "PD分析",
    "ADA分析": "ADA分析",
    "QT分析": "QT分析",
}

DESIGN_TYPE_COLS = ["SAD", "FE", "MAD", "BE", "MB"]
# 设计类型 -> Output Reference 后缀
DESIGN_TYPE_SUFFIX = {"SAD": "a", "FE": "b", "MAD": "c", "BE": "d", "MB": "e"}
# 问题3 分析物占位符（Title_CN 与 Title_EN 统一为 [Analyte]，与 TOC PH1 中一致）
PLACEHOLDER_ANALYTE = "[Analyte]"
# PK浓度/PK参数 血/尿/粪 子类型过滤：Title 中的中英文关键词
# 选(血)则排除含 尿/粪；选(尿)则排除含 血/粪；选(粪)则排除含 血/尿
SUBTYPE_TERMS = {
    "血": ["血", "Blood", "blood", "血浆", "Plasma", "plasma"],
    "尿": ["尿", "Urine", "urine"],
    "粪": ["粪", "粪便", "Feces", "feces", "Stool", "stool"],
}


def _backup_pdt(pdt_path):
    """在与 PDT 同目录下创建 archive 并备份。返回备份路径。"""
    pdt_dir = os.path.dirname(os.path.abspath(pdt_path))
    archive_dir = os.path.join(pdt_dir, "99_archive")
    os.makedirs(archive_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(pdt_path))[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{base_name}_{ts}.xlsx"
    backup_path = os.path.join(archive_dir, backup_name)
    shutil.copy2(pdt_path, backup_path)
    return backup_path


def _read_lng(setup_path):
    """
    从 setup.xlsx 的 Macro Variables sheet 中读取 LNG：
    B 列值='LNG' 时，取 C 列对应单元格的值。
    返回 LNG 字符串；若未找到返回空字符串。
    """
    wb = load_workbook(setup_path, read_only=True, data_only=True)
    if "Macro Variables" not in wb.sheetnames:
        wb.close()
        return ""
    ws = wb["Macro Variables"]
    lng_val = ""
    for row in ws.iter_rows(min_row=1, max_col=3):
        b_val = row[1].value if len(row) > 1 else None
        if b_val is not None and str(b_val).strip().upper() == "LNG":
            c_val = row[2].value if len(row) > 2 else None
            lng_val = str(c_val).strip() if c_val is not None else ""
            break
    wb.close()
    return lng_val


def _is_chinese_lng(lng_val):
    """根据 LNG 值判断是否使用中文列（Title_CN, Footnotes_CN）。"""
    if not lng_val:
        return True  # 默认中文
    v = lng_val.upper()
    if v in ("CHN", "CN", "CHINESE", "中文", "ZH", "ZH-CN"):
        return True
    return False


def _normalize_header(h):
    """规范化表头：去除空白、BOM、零宽字符、换行等"""
    if h is None:
        return ""
    s = str(h).strip().replace("\u200b", "").replace("\ufeff", "")
    return " ".join(s.split())  # 将换行、多空格归一为单空格


def _normalize_analyte_placeholder(s):
    """将 TOC 中旧占位符 <Analyte分析物> / <Analyte> 统一为 [Analyte]。"""
    if not s:
        return s
    s = str(s).replace("<Analyte分析物>", PLACEHOLDER_ANALYTE).replace("<Analyte>", PLACEHOLDER_ANALYTE)
    return s


def _read_toc_rows(toc_path):
    """
    读取 TOC 的 PH1 sheet，按列名返回行列表。
    每行为一个 dict，键为 TOC_COLS 中的列名。
    支持列别名：Footnote_CN -> Footnotes_CN, Footnote_EN -> Footnotes_EN
    """
    wb = load_workbook(toc_path, read_only=True, data_only=True)
    if "PH1" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["PH1"]
    col_idx = {}  # canonical name -> 0-based column index
    alias_map = {"Footnote_CN": "Footnotes_CN", "Footnote_EN": "Footnotes_EN"}
    rows = []
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if not col_idx:
            for i, h in enumerate(vals):
                norm = _normalize_header(h)
                if norm in TOC_COLS:
                    col_idx[norm] = i
                elif norm in alias_map:
                    col_idx[alias_map[norm]] = i
            continue
        row_dict = {}
        for col_name, idx in col_idx.items():
            val = vals[idx] if idx < len(vals) else None
            row_dict[col_name] = val
        rows.append(row_dict)
    wb.close()
    return rows


def _filter_and_expand_rows(toc_rows, design_types, endpoints, use_cn, analyte_names=None):
    """
    按基准行、终点额外添加、设计类型展开，生成 Deliverables 行列表。
    若 analyte_names 有值且 Title 含 [Analyte]，则按分析物展开；若问题2 选了 PK浓度/PK参数 且问题3 为空，则 [Analyte] 赋空值。
    返回 list of dict: {Category, Output Type, Title, Population, Footnotes, Output Reference}
    """
    use_title = "Title_CN" if use_cn else "Title_EN"
    use_footnotes = "Footnotes_CN" if use_cn else "Footnotes_EN"
    placeholder = PLACEHOLDER_ANALYTE  # Title_CN / Title_EN 统一使用 [Analyte]
    analytes = [a.strip() for a in (analyte_names or "").split("|") if a.strip()] if analyte_names else []

    # 1. 基准行：排除 EXCLUDED_CATEGORY_CN
    base_categories = set()
    for ep in endpoints:
        if ep in ENDPOINT_TO_CATEGORY:
            base_categories.add(ENDPOINT_TO_CATEGORY[ep])

    # 1.1 PK浓度/PK参数 血/尿/粪 过滤：选(血)排除 尿/粪；选(尿)排除 血/粪；选(粪)排除 血/尿
    selected_subtypes = set()
    for ep in ["PK浓度(血)", "PK浓度(尿)", "PK浓度(粪)", "PK参数(血)", "PK参数(尿)", "PK参数(粪)"]:
        if ep in endpoints:
            selected_subtypes.add(ep[-2])  # 取 "血"/"尿"/"粪"
    excluded_subtype_terms = []
    for st, terms in SUBTYPE_TERMS.items():
        if st not in selected_subtypes:
            excluded_subtype_terms.extend(terms)

    def _title_contains_excluded_subtype(row):
        """Title_CN 或 Title_EN 含未选中的 尿/粪/血 等关键词则排除"""
        if not excluded_subtype_terms or not selected_subtypes:
            return False
        title_cn = str(row.get("Title_CN") or "")
        title_en = str(row.get("Title_EN") or "")
        text = title_cn + " " + title_en
        text_lower = text.lower()
        for term in excluded_subtype_terms:
            if len(term) <= 2:  # 中文单字
                if term in title_cn or term in title_en:
                    return True
            elif term.lower() in text_lower:
                return True
        return False

    selected_rows = []
    for r in toc_rows:
        cat_cn = (r.get("Category_CN") or "").strip()
        if not cat_cn:
            continue
        if cat_cn in EXCLUDED_CATEGORY_CN:
            if cat_cn not in base_categories:
                continue
        # PK浓度/PK参数 行：排除 Title 中含未选中子类型（尿/粪等）的观测
        if cat_cn in ("PK浓度", "PK参数") and _title_contains_excluded_subtype(r):
            continue
        selected_rows.append(r)

    # 2. 对每行按 design_types 展开
    result = []
    for r in selected_rows:
        template_num = r.get("Template#")
        if template_num is None:
            template_num = ""
        else:
            template_num = str(template_num).strip()
        output_type = r.get("Output Type") or ""
        title = _normalize_analyte_placeholder(r.get(use_title) or "")
        population = r.get("Population") or ""
        footnotes = r.get(use_footnotes) or ""

        dt_cols_present = [c for c in DESIGN_TYPE_COLS if c in r and r.get(c) is not None]
        for dt in DESIGN_TYPE_COLS:
            if dt not in design_types:
                continue
            dt_val = r.get(dt)
            if dt_cols_present:
                if dt_val is None or (isinstance(dt_val, str) and not str(dt_val).strip()):
                    continue
            # 问题1 设计类型后缀：按实际选择项的序号，后缀为 "." + 序号（如选 SAD、MAD 则 SAD=1、MAD=2）
            design_ordinal = design_types.index(dt) + 1
            # Title 显示：多选时加 " - SAD" 等（中英文括号均统一为 " - "）
            if len(design_types) == 1:
                base_title = title
            else:
                base_title = f"{title} - {dt}" if title else dt

            # OUTREF 规则：先添加 [Analyte] 后缀（用 "." 连接），后添加问题1 序号后缀 "." + design_ordinal
            def _build_out_ref(analyte_idx=None):
                parts = [template_num] if (template_num and template_num.strip()) else []
                if analyte_idx is not None:
                    parts.append(str(analyte_idx))
                parts.append(str(design_ordinal))
                return ".".join(parts)

            # 问题3：若 Title 含 [Analyte] 且 analyte_names 有值则按分析物展开；否则将 [Analyte] 赋空（问题3 空值时）
            if placeholder in base_title and analytes:
                for idx, analyte in enumerate(analytes, start=1):
                    title_with_dt = base_title.replace(placeholder, analyte)
                    out_ref = _build_out_ref(analyte_idx=idx)  # 先 analyte 后缀（.idx），后 design 后缀（.design_ordinal）
                    result.append({
                        "Category": "Output",
                        "Output Type": output_type,
                        "Title": title_with_dt,
                        "Population": population,
                        "Footnotes": footnotes,
                        "Output Reference": out_ref,
                        "Validation Level": "Non-critical",
                        "Developers": "Gang Cheng",
                        "Validators": "Jianling Ren",
                        "Validated by Programmer/Statistician": "Not Started",
                    })
            else:
                # 问题2 选了 PK浓度/PK参数 且问题3 为空时，[Analyte] 赋空值
                title_with_dt = base_title.replace(placeholder, analytes[0] if analytes else "") if placeholder in base_title else base_title
                out_ref = _build_out_ref()  # 无 analyte 后缀，仅 template + "." + design_ordinal
                result.append({
                    "Category": "Output",
                    "Output Type": output_type,
                    "Title": title_with_dt,
                    "Population": population,
                    "Footnotes": footnotes,
                    "Output Reference": out_ref,
                    "Validation Level": "Non-critical",
                    "Developers": "Gang Cheng",
                    "Validators": "Jianling Ren",
                    "Validated by Programmer/Statistician": "Not Started",
                })
    return result


# TOC.xlsx 的 TOC sheet 列名（与 generate_pdt.sas 等一致）
TOC_SHEET_COLS = ["OUTTYPE", "OUTREF", "OUTTITLE", "OUTPOP", "OUTNOTE"]


def gen_toc_study(template_path, study_path, setup_path, design_types, endpoints, analyte_names=None):
    """
    根据 TOC_template.xlsx 与前三个问题（设计类型、终点、分析物），筛选并展开后生成 TOC.xlsx。
    TOC sheet 列映射：OUTTYPE<-Output Type, OUTREF<-Output Reference, OUTTITLE<-Title, OUTPOP<-Population, OUTNOTE<-Footnotes。
    """
    use_cn = True
    if setup_path and os.path.isfile(setup_path):
        lng_val = _read_lng(setup_path)
        use_cn = _is_chinese_lng(lng_val)

    toc_rows = _read_toc_rows(template_path)
    if not toc_rows:
        return False, "TOC_template 的 PH1 sheet 未找到或为空"

    new_rows = _filter_and_expand_rows(toc_rows, design_types, endpoints, use_cn, analyte_names)

    # 映射为 TOC sheet 行：OUTTYPE, OUTREF, OUTTITLE, OUTPOP, OUTNOTE
    toc_sheet_rows = []
    for r in new_rows:
        toc_sheet_rows.append({
            "OUTTYPE": r.get("Output Type") or "",
            "OUTREF": r.get("Output Reference") or "",
            "OUTTITLE": r.get("Title") or "",
            "OUTPOP": r.get("Population") or "",
            "OUTNOTE": r.get("Footnotes") or "",
        })

    d = os.path.dirname(study_path)
    if d:
        os.makedirs(d, exist_ok=True)

    # 若原 TOC.xlsx 已存在，先备份到同目录下 99_archive，文件名加年月日时分秒
    if os.path.isfile(study_path):
        study_dir = os.path.dirname(os.path.abspath(study_path))
        archive_dir = os.path.join(study_dir, "99_archive")
        os.makedirs(archive_dir, exist_ok=True)
        base_name = os.path.splitext(os.path.basename(study_path))[0]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(archive_dir, f"{base_name}_{ts}.xlsx")
        shutil.copy2(study_path, backup_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "TOC"
    for col_idx, col_name in enumerate(TOC_SHEET_COLS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(toc_sheet_rows, start=2):
        for col_idx, col_name in enumerate(TOC_SHEET_COLS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # 按内容长度适当调整列宽（中文字符按约 2 单位估算）
    def _cell_width(val):
        if val is None:
            return 0
        s = str(val)
        n = 0
        for c in s:
            n += 2 if "\u4e00" <= c <= "\u9fff" else 1
        return n
    for col_idx in range(1, len(TOC_SHEET_COLS) + 1):
        col_letter = get_column_letter(col_idx)
        max_w = 8
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            max_w = max(max_w, _cell_width(cell.value))
        # 列宽 = 内容宽度 + 余量，限制在 8~55 之间
        ws.column_dimensions[col_letter].width = min(55, max(8, max_w + 2))

    wb.save(study_path)
    wb.close()
    return True, "已生成 TOC.xlsx（TOC sheet 共 %d 行）。" % len(toc_sheet_rows)


def _find_header_row_and_cols(ws):
    """
    在 Deliverables sheet 中查找表头行和列索引。
    支持 PDT_COL_ALIASES 中的列别名。
    返回 (header_row_idx, col_name_to_idx) 或 (None, {})。
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100), start=1):
        col_name_to_idx = {}
        for col_idx, cell in enumerate(row, start=1):
            val = cell.value
            if val is None:
                continue
            norm = _normalize_header(val)
            if norm in PDT_DELIVERABLES_COLS:
                col_name_to_idx[norm] = col_idx
            elif norm in PDT_COL_ALIASES:
                col_name_to_idx[PDT_COL_ALIASES[norm]] = col_idx
        if "Category" in col_name_to_idx and len(col_name_to_idx) >= 2:
            return row_idx, col_name_to_idx
    return None, {}


# 新增行使用的默认行填充（与原有 Output 行一致的浅蓝，避免因条件格式导致新行为白底）
DEFAULT_DATA_ROW_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")

def _get_first_output_row_fill(ws, header_row, col_name_to_idx):
    """
    在删除前读取第一个 Category=Output 行的单元格填充，用于后续追加行时保持一致背景色。
    若该行无有效填充（如由条件格式控制），则返回默认浅蓝填充。
    """
    cat_col = col_name_to_idx.get("Category")
    if not cat_col:
        return DEFAULT_DATA_ROW_FILL
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=cat_col)
        if cell.value is not None and _normalize_header(cell.value) == "Output":
            # 从该行任意有填充的单元格取 fill；若整行无直接填充则用默认
            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                if c.fill and c.fill.fgColor and getattr(c.fill.fgColor, "rgb", None):
                    return copy(c.fill)
            return DEFAULT_DATA_ROW_FILL
    return DEFAULT_DATA_ROW_FILL


def _delete_output_rows(ws, header_row, col_name_to_idx):
    """从后向前删除 Category='Output' 的行（OUTCAT 等别名同义）。"""
    cat_col = col_name_to_idx.get("Category")
    if not cat_col:
        return
    rows_to_delete = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=cat_col)
        val = cell.value
        if val is not None and _normalize_header(val) == "Output":
            rows_to_delete.append(row_idx)
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)


def _find_last_data_row(ws, header_row, cat_col):
    """找到 Category 列中最后一个非空行的行号。"""
    last_row = header_row
    for row_idx in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=cat_col).value
        if val is not None and str(val).strip():
            last_row = row_idx
    return last_row


def _apply_data_validations(ws, start_row, num_rows, col_name_to_idx, list_values_sheet="List Values"):
    """对新追加的行应用数据验证。"""
    qs = quote_sheetname(list_values_sheet)
    end_row = start_row + num_rows - 1
    if end_row < start_row:
        return
    # Developers: 序列来自 List Values 的 A 列（初始值已在行数据中设为 Gang Cheng）
    dev_col = col_name_to_idx.get("Developers")
    if dev_col:
        dv = DataValidation(type="list", formula1=f"={qs}!$A$2:$A$200", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{_col_letter(dev_col)}{start_row}:{_col_letter(dev_col)}{end_row}")
    # Validators: 序列来自 List Values 的 A 列（初始值已在行数据中设为 Jianling Ren）
    val_col = col_name_to_idx.get("Validators")
    if val_col:
        dv = DataValidation(type="list", formula1=f"={qs}!$A$2:$A$200", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{_col_letter(val_col)}{start_row}:{_col_letter(val_col)}{end_row}")
    # Validation Level: $E$2:$E$4
    vl_col = col_name_to_idx.get("Validation Level")
    if vl_col:
        dv = DataValidation(type="list", formula1=f"={qs}!$E$2:$E$4", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{_col_letter(vl_col)}{start_row}:{_col_letter(vl_col)}{end_row}")
    # Output Status: $F$2:$F$3
    os_col = col_name_to_idx.get("Output Status")
    if os_col:
        dv = DataValidation(type="list", formula1=f"={qs}!$F$2:$F$3", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{_col_letter(os_col)}{start_row}:{_col_letter(os_col)}{end_row}")
    # Validated by Programmer/Statistician: $G$2:$G$4
    vbp_col = col_name_to_idx.get("Validated by Programmer/Statistician")
    if vbp_col:
        dv = DataValidation(type="list", formula1=f"={qs}!$G$2:$G$4", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{_col_letter(vbp_col)}{start_row}:{_col_letter(vbp_col)}{end_row}")


def _col_letter(col_idx):
    """列索引(1-based)转Excel列字母。"""
    return get_column_letter(col_idx)


def _append_deliverables_rows(ws, new_rows, col_name_to_idx, header_row, row_fill=None):
    """在数据区末尾（最后一行有内容的行之后）追加新行，应用数据验证，并统一新行背景色。"""
    cat_col = col_name_to_idx.get("Category")
    if cat_col:
        start_row = _find_last_data_row(ws, header_row, cat_col) + 1
    else:
        start_row = ws.max_row + 1
    fill = row_fill if row_fill is not None else DEFAULT_DATA_ROW_FILL
    max_col = ws.max_column
    for i, row_data in enumerate(new_rows):
        r = start_row + i
        for col_name, val in row_data.items():
            col_idx = col_name_to_idx.get(col_name)
            if col_idx:
                ws.cell(row=r, column=col_idx, value=val)
        for col_idx in range(1, max_col + 1):
            ws.cell(row=r, column=col_idx).fill = copy(fill)
    if new_rows:
        _apply_data_validations(ws, start_row, len(new_rows), col_name_to_idx)


def gen_pdt_deliverables(pdt_path, toc_path, setup_path, design_types, endpoints, analyte_names=None):
    """
    备份原PDT，按TOC与用户选择生成 Deliverables sheet 中 Category=Output 的行。

    Args:
        pdt_path: 项目层面 PDT.xlsx 路径
        toc_path: TOC_template.xlsx 路径
        setup_path: setup.xlsx 路径（与 PDT 同目录）
        design_types: 设计类型列表，如 ["SAD","FE","MAD"]
        endpoints: 终点列表，如 ["PK浓度(血)","PK参数(血)"]
        analyte_names: 分析物名称（可选，暂未参与逻辑）

    Returns:
        (success: bool, message: str)
    """
    try:
        # 1. 备份
        backup_path = _backup_pdt(pdt_path)

        # 2. 读 LNG
        lng_val = _read_lng(setup_path)
        use_cn = _is_chinese_lng(lng_val)

        # 3. 读 TOC PH1
        toc_rows = _read_toc_rows(toc_path)
        if not toc_rows:
            return False, "TOC PH1 未找到或为空"

        # 4. 筛选与展开
        new_rows = _filter_and_expand_rows(toc_rows, design_types, endpoints, use_cn, analyte_names)

        # 5. 写 Deliverables
        wb = load_workbook(pdt_path, data_only=False)
        if "Deliverables" not in wb.sheetnames:
            wb.close()
            return False, "PDT 中未找到 Deliverables sheet"

        ws = wb["Deliverables"]
        header_row, col_name_to_idx = _find_header_row_and_cols(ws)
        if header_row is None or "Category" not in col_name_to_idx:
            wb.close()
            return False, "Deliverables sheet 中未找到所需列（Category 等）"

        row_fill = _get_first_output_row_fill(ws, header_row, col_name_to_idx)
        _delete_output_rows(ws, header_row, col_name_to_idx)
        _append_deliverables_rows(ws, new_rows, col_name_to_idx, header_row, row_fill)

        # 强制 Excel 打开时自动重算公式（避免 #VALUE! 需手动双击刷新）
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
        if wb.calculation.calcId is not None:
            wb.calculation.calcId = (wb.calculation.calcId or 0) + 1

        wb.save(pdt_path)
        wb.close()

        msg_line1 = f"1. PDT Deliverables 表单已增加{len(new_rows)}行 TFLs记录。"
        msg_line2 = "2. 原PDT已备份至 99_archive 文件夹。"
        return True, msg_line1 + "\n" + msg_line2
    except Exception as e:
        return False, str(e)
